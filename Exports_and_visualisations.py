import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from typing import Protocol
import re

from openpyxl import Workbook

from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter


class FileWriter(Protocol):
    def write(self, df): ...

    def set_keys(self, keys): ...


class DefaultWriter:
    def set_keys(self, keys):
        return

    def write(self, df):
        raise NotImplementedError("Writer is not defined!")


class Writer:
    def __init__(
        self,
        output_file="output.xlsx",
        fill_missing_with="",
        highlight_missing=False,
        normalize_output=None,
    ):
        self.output_file = output_file
        self.fill_missing_with = fill_missing_with
        self.highlight_missing = highlight_missing
        self.keys = []
        self.templates = {}
        self.normalize_output = normalize_output

    def set_keys(self, keys):
        """
        Словарь вида
            __names: Список имён колонок
            {alias}: Список имён колонок длины len(__name), для получения значения {alias} для соответствующей колонки
        """
        if "__names" not in keys:
            raise ValueError(
                "Отсутствует информация про имена колонок. Нет записи под ключём __names."
            )
        self.keys = keys

    def set_templates(self, templates):
        formatter = re.compile(
            r"(.*?)\{(.*?)\}",
        )
        tmp = {}
        for key, template in templates.items():
            tmp[key] = []
            template_lines = template.strip().split("\n")
            for line in template_lines:
                line = line.strip()
                parsed = re.findall(formatter, line)
                extra = line.split("}")[-1]
                template_string = ""
                template_vars = []
                for elem in parsed:
                    if len(elem[1].split(":")) > 1:
                        template_string += elem[0] + "{:" + elem[1].split(":")[1] + "}"
                    else:
                        template_string += elem[0] + "{}"
                    template_vars.append(elem[1].split(":")[0])
                template_string += extra
                tmp[key].append((template_string.strip(), template_vars))

        self.templates = tmp

    def write(self, df):
        wb = Workbook()
        filled_df = df.fillna(self.fill_missing_with)

        for key in self.templates:
            wb.create_sheet(key)

        for sheet_name, template in self.templates.items():
            worksheet = wb[sheet_name]
            worksheet.append([df.index.name, *self.keys.get("__names", [])])  # Header
            for index, row in filled_df.iterrows():
                for row_n, template_row in enumerate(template):
                    idx = 0
                    if row_n == 0:
                        col_list = [index]
                    else:
                        col_list = [self.fill_missing_with]
                    for elem in self.keys.get("__names", []):
                        datas = list(
                            map(lambda x: row[self.keys[x][idx]], template_row[1])
                        )
                        flag = True
                        for item in filter(
                            lambda x: x != self.fill_missing_with, datas
                        ):
                            flag = False
                            break
                        if flag:
                            col_list.append(self.fill_missing_with)
                        else:
                            col_list.append(template_row[0].format(*datas))
                        idx += 1

                    worksheet.append(col_list)
            if self.normalize_output:
                dim_holder = DimensionHolder(worksheet=worksheet)

                for col in range(worksheet.min_column, worksheet.max_column + 1):
                    dim_holder[get_column_letter(col)] = ColumnDimension(
                        worksheet, min=col, max=col, width=self.normalize_output
                    )

                worksheet.column_dimensions = dim_holder

        try:
            wb.save(self.output_file)
            print(f"Файл успешно сохранен: {self.output_file}")
        except:
            print(f"Ошибка сохранения в файл {self.output_file}")
            raise


def dict_of_dfs_to_excel(
    df_dict, index_name="Index", writer: FileWriter = DefaultWriter(), sort_index=False
):
    """
    Преобразует словарь pandas DataFrame в Excel таблицу.

    Parameters:
    -----------
    df_dict : dict
        Словарь вида {'название_датафрейма': DataFrame, ...}
        Датафреймы должны иметь индекс.
    index_name : str
        Название столбца с индексами в результирующей таблице.
    writer:
        Объект для записи в файл.
    """
    # result_df = None
    # for name, df in df_dict.items():
    #     if result_df is None:
    #         result_df = df.add_prefix(f"{name}_")
    #     else:
    #         result_df = pd.merge(
    #             left= result_df,
    #             right= df.add_prefix(f"{name}_"),
    #             left_index= True,
    #             right_index= True,
    #             how='outer',
    #         )
    result_df = pd.concat(
        map(
            lambda x: x[1].add_prefix(f"{x[0]}_"), df_dict.items()
        ),  # Преобразование колонок фреймов в красивый вид
        axis=1,
    )
    if sort_index:
        result_df.sort_index(inplace=True)
    result_df.index.name = index_name

    keys = {}
    keys["__names"] = []
    for key, df in df_dict.items():
        for column in df.columns:
            if column not in keys:
                keys[column] = [None] * len(keys["__names"])
            keys[column].append(f"{key}_{column}")
        keys["__names"].append(key)

    # Сохраняем в Excel
    writer.set_keys(keys)
    writer.write(result_df)

    print(f"Всего строк (уникальных индексов): {len(result_df.index)}")
    print(f"Количество DataFrame: {len(df_dict)}")

    return result_df, keys


def draw_histogram_from_dict(
    data_dict,
    title="Гистограмма",
    xlabel="Категории",
    ylabel="Значения",
    hist_extra_config=None,
    figsize=(10, 6),
):
    """
    Отрисовка гистограммы из словаря {label: value}

    Args:
        data_dict: dict, словарь с метками и значениями
        title: str, заголовок графика
        xlabel: str, подпись оси X
        ylabel: str, подпись оси Y
    """
    labels = list(data_dict.keys())
    values = list(data_dict.values())

    default_hist_config = {"color": "skyblue", "edgecolor": "black"}

    if hist_extra_config is not None:
        default_hist_config.update(hist_extra_config)

    plt.figure(figsize=figsize)
    plt.bar(labels, values, **default_hist_config)

    plt.title(title, fontsize=16)
    plt.xlabel(xlabel, fontsize=12)
    plt.ylabel(ylabel, fontsize=12)

    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.grid(axis="y", alpha=0.3)

    plt.show()
